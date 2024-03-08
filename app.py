import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from log import *

data_folder = './Data'
db_name='KINERJA'
tbl_name='TIERING_AKUN'

FILE_AJ = ['PAJK Data Pendukung Desain PPP 2022.xlsx','PAJS Data Pendukung Desain PPP 2022.xlsx']
FILE_AU= ['PAUK Data Pendukung Desain PPP 2022.xlsx','PAUS Data Pendukung Desain PPP 2022.xlsx']

months  =['Januari','Februari','Maret','April','Mei','Juni','Juli','Agustus','September','Oktober','November','Desember']

products_aj=['Kematian Jangka Warsa','Endowment dan/atau Kombinasinya','Seumur Hidup','Anuitas','Kematian Ekawarsa','Kecelakaan Diri','Kesehatan','Lainnya','Jumlah']
products_au=['Harta Benda (Property)','Kendaraan Bermotor (Own Damage, Third Party Liability, dan Personal Accident)','Pengangkutan (Marine Cargo)','Rangka Kapal (Marine Hull)',
            'Rangka Pesawat (Aviation Hull)','Satelit','Energi Onshore (Oil and Gas)',
            'Energi Offshore (Oil and Gas)','Rekayasa (Engineering)',
            'Tanggung Gugat (Liability)',
            'Kecelakaan Diri','Kesehatan',
            'Kredit (Credit)','Suretyship','Aneka','Jumlah']

tierings =['0 < s.d <= Rp100.000.000','Rp100.000.000 < s.d <= Rp200.000.000',
        'Rp200.000.000 < s.d <= Rp300.000.000','Rp300.000.000 < s.d <= Rp500.000.000',
        '> Rp500.000.000']

columns=['Periode','FIN','Produk','Akun','Sektor','Tiering','Jumlah_Polis','Jumlah_Peserta','Total']

accounts_konven=['Cadangan Teknis','Cadangan Premi','CAPYBMP',
    'Cadangan Klaim','Cadangan Catastrophic','Aset Reasuransi','Utang Klaim','UP'] 

accounts_syariah=['Penyisihan Teknis','Penyisihan Kontribusi','PAKYBPM',
    'Penyisihan Klaim','Penyisihan Catastrophic','Aset Reasuransi','Utang Klaim','UP'] 

def get_report_date(year,month):
    my_month = months.index(month) + 1
    cons_tanggal ="{}-{:02}".format(year,my_month)
    final_tanggal = datetime.strptime(cons_tanggal, "%Y-%m")
    report_date = final_tanggal - relativedelta(day=31)
    return report_date.date()

def dump_to_db(db_name,table_name,df,report_date,account,sector):
    from sqlalchemy import create_engine
    from sqlalchemy.exc import SQLAlchemyError
    db = create_engine('sqlite:///{}.db'.format(db_name))
    # drop 'records table'
    try:
        with db.connect() as conn:
            query="DELETE FROM {} WHERE strftime('%Y-%m-%d',Periode)='{}' and Akun='{}' and sektor='{}'".format(table_name,report_date,account,sector)
            print(query)
            r_set=conn.exec_driver_sql(query)
    except SQLAlchemyError as e:
        error = str(e)
        print(error)
    else:
        print("No of Records deleted : ",r_set.rowcount)
    finally:
        df.to_sql(table_name, db, if_exists='append')

def get_columns(dbname,tablename):
    import pandas as pd
    import sqlite3
    # Read sqlite query results into a pandas DataFrame
    con = sqlite3.connect('{}.db'.format(dbname))
    df = pd.read_sql_query('SELECT * from {} limit 0'.format(tablename), con)
    return df

def dataProcessing(df,account,sector,isJiwa=True):
    
    df=df.iloc[11:,2:]

    df=df.dropna()

    dfFIN=pd.DataFrame(df.iloc[:,0])
    dfFIN.columns=['FIN']
    
    dfPeriode=pd.DataFrame(df.iloc[:,1])
    dfPeriode.columns=['Periode']
    
    periode = dfPeriode.iloc[0,0].strftime('%Y-%m-%d')

    #produk asuransi jiwa
    if isJiwa: # means konvensional
        df_datas= [df.iloc[:,2:19],df.iloc[:,20:37],df.iloc[:,38:55],
        df.iloc[:,56:73],df.iloc[:,74:91],df.iloc[:,92:109],
        df.iloc[:,110:127],df.iloc[:,128:145],df.iloc[:,146:163]]
    else:
        df_datas=[]
        for r in range(2,df.shape[1],18):
            df_datas.append(df.iloc[:,r:r+17])
        
    df_account = pd.DataFrame()
    
    for idx,df_data in enumerate(df_datas):
        if isJiwa: 
            produk  = products_aj[idx]
        else:
            produk  = products_au[idx]
        
        akun        = account
        
        #columns=['FIN','Periode','Produk','Akun','Sektor','Tiering','Jumlah_Polis','Jumlah_Peserta','Total']
        icol=0
        for tiering in tierings:
            df_produk_tiering=pd.DataFrame()
            df_produk_tiering['Periode']= dfPeriode['Periode']
            df_produk_tiering['FIN']= dfFIN['FIN']
            df_produk_tiering['Produk']=produk
            df_produk_tiering['Akun']=akun
            df_produk_tiering['Sektor']=sector
            df_produk_tiering['Tiering']=tiering
            df_item=df_data.iloc[:,icol:icol+3]
            df_item.columns=['Jumlah_Polis','Jumlah_Peserta','Total']
            df_produk_tiering=pd.concat([df_produk_tiering,df_item],axis=1)
            #print(df_produk_tiering)
            if df_account.shape[0]==0:
                df_account=df_produk_tiering
            else:    
                df_account=pd.concat([df_account,df_produk_tiering],axis=0)
            icol +=3

    df_account.columns=columns
    dump_to_db(db_name,tbl_name,df_account,periode,account,sector)

def read_data(filename):

    files = os.path.basename(filename).split(' ')
    sector =files[0]

    workbook = load_workbook(filename)
    isKonven='K' in sector[-1:]
    isjiwa = sector[-2:] in ['JK','JS']

    if isKonven: # means konvensional
        for index,sheet_name in enumerate(workbook.sheetnames):
            sheet = workbook[sheet_name]
            account=accounts_konven[index]
            df=pd.read_excel(filename,sheet_name=sheet.title)
            dataProcessing(df,account,sector,isjiwa)
    else: #syariah
        for index,sheet_name in enumerate(workbook.sheetnames):
            sheet = workbook[sheet_name]
            account=accounts_syariah[index]    
            df=pd.read_excel(filename,sheet_name=sheet.title)
            dataProcessing(df,account,sector,isjiwa)
def init():

    info('init')
    for file in FILE_AJ:
        filename= os.path.join(data_folder,file)
        print(filename)
        info(file)
        read_data(filename)

    for file in FILE_AU:
        filename= os.path.join(data_folder,file)
        print(filename)
        info(file)
        read_data(filename)

if __name__=="__main__":
    loginit()
    init()
